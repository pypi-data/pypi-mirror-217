from jsonpath import jsonpath
from deepdiff import DeepDiff
import base64
import os


class PytestFunctions:

    # 字典字段显示与隐藏
    def jsonKeysShow(self, data, keys, is_show=True):
        def dictHandle(data, is_show):
            d = dict()
            is_show and [d.update({k: v}) for k, v in data.items() if k in keys] or [d.update({k: v}) for k, v in data.items() if k not in keys]
            return d
        if isinstance(data, dict):
            return dictHandle(data, is_show)
        elif isinstance(data, list):
            return [isinstance(i, dict) and dictHandle(i, is_show) or i for i in data]

    # json排序
    def jsonSort(self, data):
        def dictSort(data):
            d = dict(sorted(data.items(), key=lambda x: x[0]))
            return d
        if isinstance(data, dict):
            return dictSort(data)
        elif isinstance(data, list):
            return [isinstance(item, dict) and dictSort(item) or item for item in data]
        else:
            return data

    # 读取xlsx并转为LIST_DICT
    def readXlsx(self, file_path, sheet_name='Sheet1', head=True) -> list:
        """文件路径，sheet页"""
        import openpyxl
        myxls = openpyxl.load_workbook(file_path)
        activeSheet = myxls[sheet_name]
        if head:
            keys, xlsxData = list(), list()
            for row in range(1, activeSheet.max_row + 1):
                d = dict()
                for n, column in enumerate(range(1, activeSheet.max_column + 1)):
                    data = activeSheet.cell(row=row, column=column).value
                    if len(keys) < activeSheet.max_column:
                        keys.append(data)
                    else:
                        d[keys[n]] = data
                if d:
                    xlsxData.append(d)
        else:
            xlsxData = list()
            for row in range(1, activeSheet.max_row + 1):
                xlsxData.append(
                    [activeSheet.cell(row=row, column=column).value for column in range(1, activeSheet.max_column + 1)])
        return xlsxData

    # 写入xlsx
    def writeXlsx(self, file_path, data):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        for n, row in enumerate(data):
            if isinstance(row, dict):
                if n == 0:
                    ws.append(list(row.keys()))
                ws.append(list(row.values()))
            else:
                ws.append(row)
        wb.save(file_path)

    # 读取sql语句
    def readSql(self, sql_path, parameter=None):
        with open(sql_path, 'r', encoding='utf-8')as f:
            sql = f.read()
        if isinstance(parameter, dict):
            sql = sql.format(**parameter)
        return sql

    # 日期处理
    def timeShift(self, strftime=None, **kwargs):
        '''
        支持：years, months, days, weeks, hours, minutes, seconds, microseconds
        例子： timeShift("%Y-%m-%d", days=1)
        '''
        from dateutil.relativedelta import relativedelta
        import datetime
        new_date = datetime.datetime.now() + relativedelta(**kwargs)
        if strftime:
            new_date = new_date.strftime(strftime)
        return new_date

    # 转base64图片
    def imgObjToBase64Str(self, img_obj, img_type="png", *agrs, **kwargs):
        base64_img_str = f"data:image/{img_type};base64," + base64.b64encode(img_obj).decode(*agrs, **kwargs)
        return base64_img_str

    # 数据比对
    def dataDiff(self, data1, data2, **kwargs):
        diff = DeepDiff(data1, data2, **kwargs)
        if diff:
            raise AssertionError(diff.pretty())

    # url元组
    def getUrl(self, url):
        from urllib.parse import urlparse
        return urlparse(url)

    # 字典cookies转cookiejar格式
    def toCookiejar(self, data: dict):
        import requests
        return requests.utils.cookiejar_from_dict(data)