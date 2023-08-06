from chargemaster_parsers.parsers import PalomarChargeMasterParser, ChargeMasterEntry

import tempfile
from openpyxl import Workbook
import pytest
import io
import os


@pytest.fixture
def parser():
    yield PalomarChargeMasterParser()


def test_simple_row(parser):
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="CDM")
    ws.cell(row=1, column=2, value="CDM_DESC")
    ws.cell(row=1, column=3, value="PRICE")
    ws.cell(row=2, column=1, value=473590548)
    ws.cell(row=2, column=2, value="PFC 86-4192")
    ws.cell(row=2, column=3, value="$1,034.00")

    expected_result = [
        ChargeMasterEntry(
            procedure_identifier="473590548",
            procedure_description="PFC 86-4192",
            gross_charge=1034,
        ),
    ]

    with tempfile.TemporaryDirectory() as tmp_dir:
        filename = os.path.join(tmp_dir, "palomar.xlsx")
        wb.save(filename)
        actual_result = list(
            parser.parse_artifacts(
                {PalomarChargeMasterParser.ARTIFACT_URL: open(filename, "rb")}
            )
        )
    assert sorted(expected_result) == sorted(actual_result)


def test_institution_name(parser):
    assert PalomarChargeMasterParser.institution_name == "Palomar"
    assert parser.institution_name == "Palomar"


def test_artifact_urls(parser):
    assert (
        PalomarChargeMasterParser.artifact_urls
        == PalomarChargeMasterParser.ARTIFACT_URLS
    )
    assert parser.artifact_urls == PalomarChargeMasterParser.ARTIFACT_URLS
