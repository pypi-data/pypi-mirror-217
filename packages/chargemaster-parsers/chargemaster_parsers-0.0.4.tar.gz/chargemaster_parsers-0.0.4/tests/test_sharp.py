from chargemaster_parsers.parsers import ChargeMasterEntry, SharpChargeMasterParser

import tempfile
from openpyxl import Workbook
import pytest
import io
import os

@pytest.fixture
def parser():
    yield SharpChargeMasterParser()

def test_simple_row(parser):
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="Radiology-Ultrasound")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)

    ws.cell(row=2, column=1, value="ChargeCode")
    ws.cell(row=2, column=2, value="ChargeCode Description")
    ws.cell(row=2, column=3, value="Charge")

    ws.cell(row=3, column=1, value="414300034")
    ws.cell(row=3, column=2, value="US BX BREAST INITIAL")
    ws.cell(row=3, column=3, value="$6,720.00")

    ws.cell(row=4, column=1, value="414302054")
    ws.cell(row=4, column=2, value="PERC BX LYMPH NODE")
    ws.cell(row=4, column=3, value="$4,442.00")

    expected_result = [
        ChargeMasterEntry(
            procedure_identifier = '414300034',
            location = 'Memorial',
            procedure_description = "US BX BREAST INITIAL",
            gross_charge = 6720.0,
        ),
        ChargeMasterEntry(
            procedure_identifier = '414302054',
            location = 'Memorial',
            procedure_description = "PERC BX LYMPH NODE",
            gross_charge = 4442.0,
        )
    ]

    with tempfile.TemporaryDirectory() as tmp_dir:
        filename = os.path.join(tmp_dir, "sharp.xlsx")
        wb.save(filename)
        actual_result = list(parser.parse_artifacts({"https://www.sharp.com/chargemaster/memorial/upload/SMH3082.xlsx" : open(filename, "rb")}))
    assert sorted(expected_result) == sorted(actual_result)

def test_institution_name(parser):
    assert SharpChargeMasterParser.institution_name == "Sharp"
    assert parser.institution_name == "Sharp"

def test_artifact_urls(parser):
    assert SharpChargeMasterParser.artifact_urls == SharpChargeMasterParser.ARTIFACT_URLS
    assert parser.artifact_urls == SharpChargeMasterParser.ARTIFACT_URLS