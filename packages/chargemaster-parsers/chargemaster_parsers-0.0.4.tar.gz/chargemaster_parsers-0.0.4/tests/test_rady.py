from chargemaster_parsers.parsers import RadyChargeMasterParser, ChargeMasterEntry

import tempfile
from openpyxl import Workbook
import pytest
import io
import os

@pytest.fixture
def parser():
    yield RadyChargeMasterParser()

def test_simple_row(parser):
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="Procedure Name")
    ws.cell(row=1, column=2, value="Price")
    ws.cell(row=2, column=1, value="RCH PEDIATRIC PRIVATE ROOM CHARGE")
    ws.cell(row=2, column=2, value=8400)
    ws.cell(row=3, column=1, value="RCH PEDIATRIC SEMIPRIVATE ROOM CHG")
    ws.cell(row=3, column=2, value=8400)


    expected_result = [
        ChargeMasterEntry(
            procedure_identifier = 1,
            procedure_description = "PEDIATRIC PRIVATE ROOM CHARGE",
            gross_charge = 8400,
        ),
        ChargeMasterEntry(
            procedure_identifier = 2,
            procedure_description = "PEDIATRIC SEMIPRIVATE ROOM CHG",
            gross_charge = 8400,
        )
    ]

    with tempfile.TemporaryDirectory() as tmp_dir:
        filename = os.path.join(tmp_dir, "rady.xlsx")
        wb.save(filename)
        actual_result = list(parser.parse_artifacts({RadyChargeMasterParser.ARTIFACT_URL : open(filename, "rb")}))
    assert sorted(expected_result) == sorted(actual_result)

def test_institution_name(parser):
    assert RadyChargeMasterParser.institution_name == "Rady"
    assert parser.institution_name == "Rady"

def test_artifact_urls(parser):
    assert RadyChargeMasterParser.artifact_urls == RadyChargeMasterParser.ARTIFACT_URLS
    assert parser.artifact_urls == RadyChargeMasterParser.ARTIFACT_URLS