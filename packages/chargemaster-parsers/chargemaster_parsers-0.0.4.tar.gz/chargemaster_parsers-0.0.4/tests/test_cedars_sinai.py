from chargemaster_parsers.parsers.cedars_sinai import CedarsSinaiChargeMasterParser
from chargemaster_parsers.parsers import ChargeMasterEntry

import tempfile
from openpyxl import Workbook
import os
import pprint
import pytest
import io

@pytest.fixture
def parser():
    yield CedarsSinaiChargeMasterParser()

def test_simple_row(parser):
    wb = Workbook()
    ws = wb.active

    ws.cell(row=5, column=1, value="EAP PROC CODE")
    ws.cell(row=5, column=2, value="EAP PROC NAME")
    ws.cell(row=5, column=3, value="DEFAULT CPT/ HCPCS CODE")
    ws.cell(row=5, column=4, value="DEFAULT OP FEE SCHEDULE")
    ws.cell(row=5, column=5, value="IP/ED FEE SCHEDULE")

    ws.cell(row=6, column=1, value=2600001)
    ws.cell(row=6, column=2, value="HB IV INFUS HYDRATION 31-60 MIN")
    ws.cell(row=6, column=3, value=96360)
    ws.cell(row=6, column=4, value="$2161.71")
    ws.cell(row=6, column=5, value="$2811.66")

    ws.cell(row=7, column=1, value=2600017)
    ws.cell(row=7, column=2, value="HB APPL ON-BODY INJECTOR SUBQ INJ")
    ws.cell(row=7, column=3, value=96377)
    ws.cell(row=7, column=4, value="$844.45")
    ws.cell(row=7, column=5, value=None)

    ws.cell(row=8, column=1, value=2700005)
    ws.cell(row=8, column=2, value="HB CYTOTOXIC SPILL KIT")
    ws.cell(row=8, column=3, value=None)
    ws.cell(row=8, column=4, value="$254.09")
    ws.cell(row=8, column=5, value=None)

    expected_result = [
        ChargeMasterEntry(
            procedure_identifier = 2600001,
            location = 'all',
            procedure_description = "HB IV INFUS HYDRATION 31-60 MIN",
            gross_charge = 2161.71,
            in_patient = False
        ),
        ChargeMasterEntry(
            procedure_identifier = 2600001,
            location = 'all',
            procedure_description = "HB IV INFUS HYDRATION 31-60 MIN",
            gross_charge = 2811.66,
            in_patient = True,
            cpt_code = 96360,
        ),
        ChargeMasterEntry(
            procedure_identifier = 2600017,
            location = 'all',
            procedure_description = "HB APPL ON-BODY INJECTOR SUBQ INJ",
            gross_charge = 844.45,
            in_patient = False,
        ),
        ChargeMasterEntry(
            procedure_identifier = 2600017,
            location = 'all',
            procedure_description = "HB APPL ON-BODY INJECTOR SUBQ INJ",
            gross_charge = 844.45,
            in_patient = True,
            cpt_code = 96377,
        ),
        ChargeMasterEntry(
            procedure_identifier = 2700005,
            location = 'all',
            procedure_description = "HB CYTOTOXIC SPILL KIT",
            gross_charge = 254.09,
            in_patient = False,
        ),
    ]

    with tempfile.TemporaryDirectory() as tmp_dir:
        filename = os.path.join(tmp_dir, "cedars.xlsx")
        wb.save(filename)
        actual_result = list(parser.parse_artifacts({"https://www.cedars-sinai.org/content/dam/cedars-sinai/billing-insurance/documents/cedars-sinai-changemaster-july-2022.xlsx" : open(filename, "rb")}))
    # pprint.pprint(actual_result)
    assert sorted(expected_result) == sorted(actual_result)

def test_institution_name(parser):
    assert CedarsSinaiChargeMasterParser.institution_name == "Cedars-Sinai"
    assert parser.institution_name == "Cedars-Sinai"

def test_artifact_urls(parser):
    assert CedarsSinaiChargeMasterParser.artifact_urls == CedarsSinaiChargeMasterParser.ARTIFACT_URLS
    assert parser.artifact_urls == CedarsSinaiChargeMasterParser.ARTIFACT_URLS